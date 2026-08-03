import streamlit as st
import pandas as pd
import os
import re
import io
from datetime import datetime, timedelta, timezone
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. 유틸리티 함수 ---
def format_unit(unit, count, force_to_pkg=False):
    u_str = str(unit).upper() if pd.notna(unit) else "PKG"
    m = {'PK':'PKG', 'PL':'PLT', 'CT':'CTN'}
    base = 'PKG' if (force_to_pkg and u_str == 'PL') else m.get(u_str, u_str)
    if u_str in ['PK', 'PL', 'CT'] and count > 1: return base + 'S'
    return base

def format_number(v):
    try:
        val = float(v)
        t = f"{round(val, 3):.3f}"
        return t.rstrip('0').rstrip('.') if '.' in t else t
    except: return str(v)

def log_uploaded_filename(fn, category="SR"):
    p = "upload_log.txt"
    kst = timezone(timedelta(hours=9))
    now = datetime.now(kst).strftime("%Y-%m-%d %H:%M:%S")
    entry = f"[{now}] ({category}) {fn}\n"
    with open(p, "a", encoding='utf-8') as f:
        f.write(entry)

# [CEVA 전용] 단위 포맷 함수
def format_unit_ceva(unit, count):
    if not unit: return ""
    u = str(unit).upper().strip()
    mapping = {'PLT': 'PLT', 'PALLET': 'PLT', 'PLTS': 'PLT', 'PKG': 'PKG', 'PKGS': 'PKG', 'CTN': 'CTN', 'CTNS': 'CTN'}
    base = mapping.get(u, u)
    if count > 1:
        return base + "S"
    return base

# [CEVA 전용] 중량 포맷 함수
def format_wgt_ceva(v):
    try:
        val = float(v)
        if val == int(val):
            return str(int(val))
        return str(val)
    except:
        return str(v)

# [IST CONSOL 전용] 날짜 포맷 함수 (예: YYYY-MM-DD)
def format_date_ist(v):
    if pd.isna(v) or not v:
        return ""
    try:
        if isinstance(v, (datetime, pd.Timestamp)):
            dt = v
        else:
            v_str = str(v).strip().split(' ')[0]
            dt = pd.to_datetime(v_str)
        return dt.strftime("%Y-%m-%d")
    except:
        return str(v).split(' ')[0]

# [IST CONSOL 전용] 회사명 정제 파서
def clean_company_name(text, is_pus=False, is_shipper=True):
    if pd.isna(text) or not str(text).strip():
        return ""
    
    raw = str(text).strip()
    lines = [l.strip() for l in raw.split('\n') if l.strip()]
    if not lines:
        return ""
    
    if is_pus:
        target_lines = lines[:1]
    else:
        if is_shipper:
            if lines[0].upper().startswith("NIPPON"):
                target_lines = lines
            else:
                target_lines = lines[1:] if len(lines) > 1 else lines
        else:
            target_lines = lines
            
    if not target_lines:
        return ""
        
    full_text = " ".join(target_lines)
    
    prefixes = [
        r"^O/B\s*", r"^O/B:\s*", r"^AS\s+AGENT\s+OF\s*", r"^ON\s+BEHALF\s+OF\s*",
        r"^OB\s*", r"^OB:\s*", r"^AS\s+AGENTS\s+FOR\s*", r"^AS\s+AGENT\s+FOR\s*"
    ]
    for ptn in prefixes:
        full_text = re.sub(ptn, "", full_text, flags=re.IGNORECASE).strip()
        
    suffixes = [
        r"LTD\.", r"LIMITED", r"A\.S", r"A\.S\.", r"\bAS\b", r"INC\.",
        r"STI\.", r"STI", r"CO\.,\s*LTD\.", r"CORP\.", r"CORPORATION"
    ]
    
    words = full_text.split()
    matched_idx = -1
    for i, w in enumerate(words):
        w_clean = re.sub(r'[^A-Za-z\.]', '', w).upper()
        for sfx in suffixes:
            s_clean = re.sub(r'[^A-Za-z\.]', '', sfx).upper()
            if w_clean == s_clean or w_clean.endswith(s_clean):
                matched_idx = i
                break
        if matched_idx != -1:
            break
            
    if matched_idx != -1:
        comp_name = " ".join(words[:matched_idx+1])
    else:
        comp_name = target_lines[0]
        
    for ptn in prefixes:
        comp_name = re.sub(ptn, "", comp_name, flags=re.IGNORECASE).strip()
        
    return comp_name

# --- POD 정의 ---
POD_LIST = [
    ("전체", "ALL"),
    ("벨기에", "BEANR"), ("독일", "DEHAM"), ("덴마크", "DKAAR"), ("스페인", "ESBCN"),
    ("프랑스", "FRFOS"), ("프랑스", "FRLEH"), ("영국", "GBSOU"), ("이탈리아", "ITGOA"),
    ("네덜란드", "NLRTM"), ("노르웨이", "NOOSL"), ("폴란드", "PLGDN"), ("루마니아", "ROCND"),
    ("스웨덴", "SEGOT"), ("슬로베니아", "SIKOP"), ("터키", "TRIST")
]
POD_OPTIONS = [f"{country} ({code})" if code != "ALL" else "전체 (ALL)" for country, code in POD_LIST]

# --- 2. 페이지 설정 ---
st.set_page_config(page_title="Europe Docs tool (Cargo Tool 6)", layout="wide")
st.title("🚢 Europe Docs tool")

# 탭 생성
tab1, tab_ceva, tab_ist, tab_history, tab2 = st.tabs(["SR 정정", "CEVA(LEH)", "IST CONSOL", "선적이력", "업로드 기록"])

# ==========================================
# TAB 1: SR 정정 (Cargo Tool 6 - 대원칙 보존)
# ==========================================
with tab1:
    col_up1, col_up2, col_opt = st.columns([1.0, 1.5, 0.8])
    with col_up1:
        sr_file = st.file_uploader("1. SR 엑셀 파일 입력", type=["xlsx"], key="sr_main")
    with col_up2:
        item_file = st.file_uploader("2. 하우스리스트 → S/R NO 검색 → 엑셀내려받기 파일 입력(품목, HS CODE 추가 가능)", type=["xlsx"], key="item_sub")
    with col_opt:
        st.write("") 
        st.write("") 
        force_to_pkg = st.checkbox("코스코 PLT -> PKG 변환", value=False)
        mark_spacing = st.checkbox("MARK 란 간격 띄우기", value=False)

    st.divider()

    if sr_file:
        try:
            log_uploaded_filename(sr_file.name, "SR")
            sr_df = pd.read_excel(sr_file)
            item_dict = {}; warning_messages = []

            if "House B/L No" in sr_df.columns and "단위" in sr_df.columns:
                for _, row in sr_df.iterrows():
                    h_no_sr = str(row["House B/L No"]).strip()
                    unit_sr = str(row["단위"]).strip().upper() if pd.notna(row["단위"]) else ""
                    if h_no_sr and h_no_sr != "nan":
                        if unit_sr == "GT":
                            warning_messages.append(f"⚠️ {h_no_sr}: 단위가 GT 입니다.")

            if item_file:
                log_uploaded_filename(item_file.name, "ITEM")
                item_df = pd.read_excel(item_file, header=1)
                item_df.columns = [str(c).strip() for c in item_df.columns]
                
                if "House B/L No" in item_df.columns and "품목" in item_df.columns:
                    for _, row in item_df.iterrows():
                        h_no = str(row["House B/L No"]).strip()
                        raw_desc = str(row["품목"]).strip() if pd.notna(row["품목"]) else ""
                        
                        if h_no and h_no != "nan":
                            all_lines = [l.strip() for l in raw_desc.split('\n') if l.strip()]
                            found_hs_list = []
                            for line in all_lines:
                                if re.match(r'^[0-9.]{4,11}$', line):
                                    found_hs_list.append(line)
                            
                            detected_hs = found_hs_list[-1] if found_hs_list else ""
                            detected_desc_pure = raw_desc
                            if detected_hs:
                                detected_desc_pure = raw_desc.replace(detected_hs, "").strip()
                            
                            item_dict[h_no] = {"desc": raw_desc, "hs": detected_hs}
                            
                            has_multiple = False
                            if len(all_lines) >= 3:
                                for i in range(len(all_lines) - 2):
                                    if re.match(r'^[0-9.]{4,10}$', all_lines[i]) and not re.match(r'^[0-9.]{4,10}$', all_lines[i+1]):
                                        has_multiple = True
                                        break
                            if has_multiple:
                                warning_messages.append(f"📢 {h_no}: 다중 품목 -> 수기로 컨테이너 별 품목을 나눠주세요ㅎㅎ")

                            is_desc_empty = not detected_desc_pure or detected_desc_pure.lower() == "nan" or detected_desc_pure.strip() == ""
                            is_hs_empty = not detected_hs or detected_hs.strip() == ""

                            if is_desc_empty and is_hs_empty:
                                warning_messages.append(f"⚠️ {h_no}: 품목, HS CODE 가 공란입니다!")
                            elif is_desc_empty:
                                warning_messages.append(f"⚠️ {h_no}: 품목이 공란입니다!")
                            elif is_hs_empty:
                                warning_messages.append(f"⚠️ {h_no}: HS CODE 가 공란입니다!")
                            else:
                                clean_hs_digits = re.sub(r'[^0-9]', '', detected_hs)
                                if "." in detected_hs:
                                    if not re.match(r'^\d{4}\.\d{2}$', detected_hs):
                                        warning_messages.append(f"⚠️ {h_no}: HS CODE 형식 오류")
                                elif len(clean_hs_digits) != 6:
                                    warning_messages.append(f"⚠️ {h_no}: HS CODE 형식 오류")
                            
                            if "MAGNET" in raw_desc.upper():
                                warning_messages.append(f"⚠️ {h_no}: 자성물질 MSDS 필요!")
                            
                            if detected_hs:
                                clean_hs = str(detected_hs).replace(".", "").replace(" ", "")
                                if clean_hs == "242400":
                                    warning_messages.append(f"⚠️ {h_no}: 유효하지 않은 HS CODE / HOUSEHOLD GOODS 는 9905.00 을 써주세요.")

            cols = ['House B/L No', '컨테이너 번호', 'Seal#1', '포장갯수', '단위', 'Weight', 'Measure']
            df = sr_df[cols].copy().dropna(subset=['House B/L No'])
            df['Seal#1'] = df['Seal#1'].fillna('').astype(str).str.split('.').str[0]
            df['단위'] = df['단위'].fillna('PKG')
            
            total = df.groupby(['컨테이너 번호', 'Seal#1']).agg(포장갯수=('포장갯수','sum'), Weight=('Weight','sum'), Measure=('Measure','sum')).reset_index()
            marks = df.groupby(['컨테이너 번호', 'Seal#1'])['House B/L No'].unique().reset_index()
            desc_df = df.sort_values(['컨테이너 번호', 'Seal#1', 'House B/L No'])
            
            lines = []
            num_containers = len(total)
            if num_containers > 1:
                g_p = int(total['포장갯수'].sum())
                total_line = f"TOTAL: {g_p} PKGS / {format_number(total['Weight'].sum())} KGS / {format_number(total['Measure'].sum())} CBM"
                lines.extend(["[GRAND TOTAL]", total_line, "-" * (len(total_line) + 10)]) 
            
            for _, r in total.iterrows():
                lines.append(""); lines.append(f"{r['컨테이너 번호']} / {r['Seal#1']}")
                lines.append(f"TOTAL: {int(r['포장갯수'])} PKGS / {format_number(r['Weight'])} KGS / {format_number(r['Measure'])} CBM")
            
            lines.extend(["", "", "<MARK>", ""]) 
            for i, r in marks.iterrows():
                if i > 0: lines.append("") 
                if num_containers > 1:
                    lines.append(f"{r['컨테이너 번호']} / {r['Seal#1']}")
                    lines.append("") 
                for hbl in sorted(r['House B/L No']):
                    lines.append(hbl)
                    if num_containers <= 4 and mark_spacing: lines.append("") 
                if not (num_containers <= 4 and mark_spacing): lines.append("") 
            
            lines.extend(["", "<DESCRIPTION>", ""]) 
            prev = (None, None)
            for _, r in desc_df.iterrows():
                cur = (r['컨테이너 번호'], r['Seal#1'])
                if cur != prev:
                    if prev[0] is not None: lines.extend(["", ""]) 
                    if num_containers > 1: lines.extend([f"{cur[0]} / {cur[1]}", ""])
                    prev = cur
                h_no_raw = str(r['House B/L No']).strip()
                lines.append(h_no_raw)
                lines.append(f"{int(r['포장갯수'])} {format_unit(r['단위'], r['포장갯수'], force_to_pkg)} / {format_number(r['Weight'])} KGS / {format_number(r['Measure'])} CBM")
                if h_no_raw in item_dict:
                    info = item_dict[h_no_raw]
                    if info["desc"] and info["desc"].lower() != "nan": lines.append(info["desc"])
                lines.append("")
            
            result = "\n".join(lines)
            res_head, res_down = st.columns([3, 1])
            with res_head: st.subheader("정리 결과")
            with res_down: st.download_button("💾 메모장 다운로드", result, f"SR_{sr_file.name.split('.')[0]}.txt", use_container_width=True)
            
            if warning_messages:
                combined_warning = "\n".join(warning_messages)
                st.markdown(f'<div style="display:inline-block;padding:5px 15px;border-radius:5px;background-color:rgba(255, 75, 75, 0.1);border:1px solid rgb(255, 75, 75);color:rgb(255, 75, 75);font-family:sans-serif;font-size:14px;line-height:1.5;white-space:pre-wrap;margin-bottom:5px;">{combined_warning}</div><br>', unsafe_allow_html=True)
            
            st.text_area("결과창", result, height=800, label_visibility="collapsed")
        except Exception as e: st.error(f"오류 발생: {e}")

# ==========================================
# TAB 2: CEVA(LEH)
# ==========================================
with tab_ceva:
    col_ceva_up = st.columns([1])[0]
    with col_ceva_up:
        ceva_file = st.file_uploader("CEVA 엑셀 파일을 업로드하세요", type=["xlsx"], key="ceva_up")
    
    if ceva_file:
        try:
            c_df = pd.read_excel(ceva_file, header=None)
            def get_val(r, c):
                try: 
                    v = c_df.iloc[r, c]
                    return str(v).strip() if pd.notna(v) else ""
                except: return ""
            
            sets = [
                {"qty": (35,8), "unit": (35,14), "wgt": (36,8), "cbm": (37,8), "hc": (38,4), "mark": (36,16), "desc": (36,34)},
                {"qty": (44,8), "unit": (44,14), "wgt": (45,8), "cbm": (46,8), "hc": (47,4), "mark": (45,16), "desc": (45,34)},
                {"qty": (58,8), "unit": (58,14), "wgt": (59,8), "cbm": (60,8), "hc": (61,4), "mark": (59,16), "desc": (59,34)},
                {"qty": (67,8), "unit": (67,14), "wgt": (68,8), "cbm": (69,8), "hc": (70,4), "mark": (68,16), "desc": (68,34)},
                {"qty": (76,8), "unit": (76,14), "wgt": (77,8), "cbm": (78,8), "hc": (79,4), "mark": (77,16), "desc": (77,34)},
                {"qty": (85,8), "unit": (85,14), "wgt": (86,8), "cbm": (87,8), "hc": (88,4), "mark": (86,16), "desc": (86,34)},
                {"qty": (94,8), "unit": (94,14), "wgt": (95,8), "cbm": (96,8), "hc": (97,4), "mark": (95,16), "desc": (95,34)}
            ]
            
            mark_lines, desc_lines = [], []
            for s in sets:
                qty_val = get_val(*s["qty"])
                if not qty_val: continue
                qty_int = int(float(qty_val)) if qty_val.replace('.','').isdigit() else 0
                unit_str = format_unit_ceva(get_val(*s["unit"]), qty_int)
                wgt_str = format_wgt_ceva(get_val(*s["wgt"]))
                hc_val_raw, mark_str, desc_str = get_val(*s["hc"]), get_val(*s["mark"]), get_val(*s["desc"])
                
                mark_lines.extend([mark_str, "", ""])
                desc_lines.append(desc_str)
                desc_lines.append(f"{qty_int} {unit_str} / {wgt_str} KGS / CBM")
                if hc_val_raw:
                    desc_lines.append(f"HC: {hc_val_raw.replace('HC:', '').strip()}")
                desc_lines.extend(["", ""]) 
            
            st.divider()
            res_col1, res_col2 = st.columns(2)
            with res_col1:
                st.subheader("<MARK>")
                st.text_area("MARK 결과", "\n".join(mark_lines), height=600, label_visibility="collapsed")
            with res_col2:
                st.subheader("<DESCRIPTION>")
                st.text_area("DESC 결과", "\n".join(desc_lines), height=600, label_visibility="collapsed")
        except Exception as e: st.error(f"오류 발생: {e}")

# ==========================================
# TAB 3: IST CONSOL (컨테이너별 자동 정렬 포함)
# ==========================================
with tab_ist:
    col_ist_up = st.columns([1.2, 1])[0]
    with col_ist_up:
        ist_file = st.file_uploader("IST 엑셀 파일을 업로드하세요", type=["xlsx"], key="ist_up")
        
    if ist_file:
        try:
            log_uploaded_filename(ist_file.name, "IST")
            raw_ist_df = pd.read_excel(ist_file)
            
            header_idx = None
            if "House B/L No" in raw_ist_df.columns:
                df1 = raw_ist_df
            else:
                for idx in range(min(5, len(raw_ist_df))):
                    row_vals = [str(v).strip() for v in raw_ist_df.iloc[idx].values]
                    if any("House B/L" in v for v in row_vals):
                        header_idx = idx
                        break
                if header_idx is not None:
                    df1 = pd.read_excel(ist_file, header=header_idx + 1)
                else:
                    df1 = raw_ist_df

            df1.columns = [str(c).strip() for c in df1.columns]
            
            hbl_col = next((c for c in df1.columns if "House B/L" in c), None)
            vessel_col = next((c for c in df1.columns if "Vessel" in c), None)
            voyage_col = next((c for c in df1.columns if "항차" in c), None)
            etd_col = next((c for c in df1.columns if "ETD" in c), None)
            eta_col = next((c for c in df1.columns if "ETA" in c), None)
            mbl_col = next((c for c in df1.columns if "Master B/L" in c), None)
            shipper_col = next((c for c in df1.columns if "Shipper" in c and "Real" not in c), None)
            consignee_col = next((c for c in df1.columns if "Consignee" in c), None)
            notify_col = next((c for c in df1.columns if "Notify" in c), None)
            weight_col = next((c for c in df1.columns if "Weight" in c), None)
            pkg_col = next((c for c in df1.columns if "포장갯수" in c), None)
            measure_col = next((c for c in df1.columns if "Measure" in c), None)
            cntr_col = next((c for c in df1.columns if "컨테이너" in c), None)
            seal_col = next((c for c in df1.columns if "Seal No" in c or "Seal#1" in c), None)

            if hbl_col:
                valid_df = df1.dropna(subset=[hbl_col]).copy()

                # ⭐ 컨테이너 번호 기준 자동 정렬 (컨테이너별 순서 배치) ⭐
                if cntr_col:
                    valid_df = valid_df.sort_values(by=[cntr_col, hbl_col], ascending=[True, True])

                first_row = valid_df.iloc[0] if len(valid_df) > 0 else None
                
                vessel_val = str(first_row[vessel_col]).strip() if (first_row is not None and vessel_col and pd.notna(first_row[vessel_col])) else ""
                voyage_val = str(first_row[voyage_col]).strip() if (first_row is not None and voyage_col and pd.notna(first_row[voyage_col])) else ""
                etd_val = format_date_ist(first_row[etd_col]) if (first_row is not None and etd_col) else ""
                eta_val = format_date_ist(first_row[eta_col]) if (first_row is not None and eta_col) else ""
                mbl_val = str(first_row[mbl_col]).strip() if (first_row is not None and mbl_col and pd.notna(first_row[mbl_col])) else ""

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "LOADING LIST"

                font_calibri_bold = Font(name="Calibri", size=11, bold=True)
                font_calibri_regular = Font(name="Calibri", size=11, bold=False)
                
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")

                thin_side = Side(border_style="thin", color="000000")
                med_side = Side(border_style="medium", color="000000")

                b_top_h = Border(top=med_side, bottom=thin_side, left=thin_side, right=thin_side)
                b_top_d = Border(top=thin_side, bottom=med_side, left=thin_side, right=thin_side)

                ws["E1"] = "Vessel"; ws["E1"].font = font_calibri_bold; ws["E1"].alignment = align_center; ws["E1"].border = Border(top=med_side, bottom=thin_side, left=med_side, right=thin_side)
                ws["F1"] = "Voyage"; ws["F1"].font = font_calibri_bold; ws["F1"].alignment = align_center; ws["F1"].border = b_top_h
                ws["G1"] = "Cut Off"; ws["G1"].font = font_calibri_bold; ws["G1"].alignment = align_center; ws["G1"].border = b_top_h
                ws["H1"] = "ETD"; ws["H1"].font = font_calibri_bold; ws["H1"].alignment = align_center; ws["H1"].border = b_top_h
                ws["I1"] = "ETA"; ws["I1"].font = font_calibri_bold; ws["I1"].alignment = align_center; ws["I1"].border = Border(top=med_side, bottom=thin_side, left=thin_side, right=med_side)
                
                ws["J1"] = "Master No"; ws["J1"].font = font_calibri_bold; ws["J1"].alignment = align_center
                ws.merge_cells("J1:L1")
                for col_idx in [10, 11, 12]:
                    ws.cell(row=1, column=col_idx).border = Border(top=med_side, bottom=thin_side)
                ws["J1"].border = Border(top=med_side, bottom=thin_side, left=med_side)
                ws["L1"].border = Border(top=med_side, bottom=thin_side, right=med_side)

                ws["M1"] = "Carrier"; ws["M1"].font = font_calibri_bold; ws["M1"].alignment = align_center; ws["M1"].border = Border(top=med_side, bottom=thin_side, left=med_side, right=med_side)

                ws["E2"] = vessel_val; ws["E2"].font = font_calibri_regular; ws["E2"].alignment = align_center; ws["E2"].border = Border(top=thin_side, bottom=med_side, left=med_side, right=thin_side)
                ws["F2"] = voyage_val; ws["F2"].font = font_calibri_regular; ws["F2"].alignment = align_center; ws["F2"].border = b_top_d
                ws["G2"] = ""; ws["G2"].font = font_calibri_regular; ws["G2"].alignment = align_center; ws["G2"].border = b_top_d
                ws["H2"] = etd_val; ws["H2"].font = font_calibri_regular; ws["H2"].alignment = align_center; ws["H2"].border = b_top_d
                ws["I2"] = eta_val; ws["I2"].font = font_calibri_regular; ws["I2"].alignment = align_center; ws["I2"].border = Border(top=thin_side, bottom=med_side, left=thin_side, right=med_side)
                
                ws["J2"] = mbl_val; ws["J2"].font = font_calibri_regular; ws["J2"].alignment = align_center
                ws.merge_cells("J2:L2")
                for col_idx in [10, 11, 12]:
                    ws.cell(row=2, column=col_idx).border = Border(top=thin_side, bottom=med_side)
                ws["J2"].border = Border(top=thin_side, bottom=med_side, left=med_side)
                ws["L2"].border = Border(top=thin_side, bottom=med_side, right=med_side)

                ws["M2"] = "MSC"; ws["M2"].font = font_calibri_regular; ws["M2"].alignment = align_center; ws["M2"].border = Border(top=thin_side, bottom=med_side, left=med_side, right=med_side)

                ws["A3"] = "POL"; ws["A3"].font = font_calibri_bold
                ws["B3"] = "BUSAN "; ws["B3"].font = font_calibri_bold
                ws["A4"] = "POD"; ws["A4"].font = font_calibri_bold
                ws["B4"] = "ISTANBUL "; ws["B4"].font = font_calibri_bold

                tbl_h = [
                    ("A6", "Shpt"), ("B6", "HBL No."), ("C6", "CNTR SIZE"), ("D6", "Shipper"),
                    ("E6", "Consignee"), ("F6", "KGS"), ("G6", "PKG'S"), ("I6", "CBM"),
                    ("J6", "R/O"), ("K6", "Conatainer No."), ("L6", "Seal No."), ("M6", "REMARKS")
                ]
                for cell_ref, title in tbl_h:
                    cell = ws[cell_ref]
                    cell.value = title
                    cell.font = font_calibri_bold
                    cell.alignment = align_center
                    cell.border = Border(top=med_side, bottom=med_side, left=thin_side, right=thin_side)

                ws.merge_cells("G6:H6")
                ws["G6"].border = Border(top=med_side, bottom=med_side, left=thin_side, right=thin_side)
                ws["H6"].border = Border(top=med_side, bottom=med_side, left=thin_side, right=thin_side)

                start_row = 7
                total_data_count = len(valid_df)

                sum_kgs = 0.0
                sum_pkgs = 0
                sum_cbm = 0.0

                b_data_cell = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

                for row_idx_0, (_, r) in enumerate(valid_df.iterrows()):
                    curr_row = start_row + row_idx_0
                    hbl_val = str(r[hbl_col]).strip() if pd.notna(r[hbl_col]) else ""
                    is_pus = hbl_val.upper().startswith("PUS")

                    raw_shipper = r[shipper_col] if shipper_col else ""
                    shipper_clean = clean_company_name(raw_shipper, is_pus=is_pus, is_shipper=True)

                    if is_pus:
                        raw_consignee = r[consignee_col] if consignee_col else ""
                        consignee_clean = clean_company_name(raw_consignee, is_pus=True, is_shipper=False)
                    else:
                        raw_notify = r[notify_col] if notify_col else ""
                        consignee_clean = clean_company_name(raw_notify, is_pus=False, is_shipper=False)

                    wgt_raw = float(r[weight_col]) if (weight_col and pd.notna(r[weight_col]) and str(r[weight_col]).replace('.','').isdigit()) else 0.0
                    pkg_raw = int(float(r[pkg_col])) if (pkg_col and pd.notna(r[pkg_col]) and str(r[pkg_col]).replace('.','').isdigit()) else 0
                    cbm_raw = float(r[measure_col]) if (measure_col and pd.notna(r[measure_col]) and str(r[measure_col]).replace('.','').isdigit()) else 0.0

                    cbm_final = 1.0 if (0 < cbm_raw < 1) else cbm_raw

                    sum_kgs += wgt_raw
                    sum_pkgs += pkg_raw
                    sum_cbm += cbm_final

                    cntr_v = str(r[cntr_col]).strip() if (cntr_col and pd.notna(r[cntr_col])) else ""
                    seal_v = str(r[seal_col]).strip().split('.')[0] if (seal_col and pd.notna(r[seal_col])) else ""

                    ws[f"A{curr_row}"] = row_idx_0 + 1; ws[f"A{curr_row}"].alignment = align_center
                    ws[f"B{curr_row}"] = hbl_val; ws[f"B{curr_row}"].alignment = align_center
                    ws[f"C{curr_row}"] = "40'HC"; ws[f"C{curr_row}"].alignment = align_center
                    ws[f"D{curr_row}"] = shipper_clean; ws[f"D{curr_row}"].alignment = align_left
                    ws[f"E{curr_row}"] = consignee_clean; ws[f"E{curr_row}"].alignment = align_left
                    ws[f"F{curr_row}"] = wgt_raw; ws[f"F{curr_row}"].alignment = align_center
                    ws[f"G{curr_row}"] = pkg_raw; ws[f"G{curr_row}"].alignment = align_center
                    ws[f"H{curr_row}"] = "PKG'S"; ws[f"H{curr_row}"].alignment = align_center
                    ws[f"I{curr_row}"] = cbm_final; ws[f"I{curr_row}"].alignment = align_center
                    ws[f"J{curr_row}"] = ""; ws[f"J{curr_row}"].alignment = align_center
                    ws[f"K{curr_row}"] = cntr_v; ws[f"K{curr_row}"].alignment = align_center
                    ws[f"L{curr_row}"] = seal_v; ws[f"L{curr_row}"].alignment = align_center
                    ws[f"M{curr_row}"] = ""; ws[f"M{curr_row}"].alignment = align_center

                    for col_l in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
                        ws[f"{col_l}{curr_row}"].font = font_calibri_regular
                        ws[f"{col_l}{curr_row}"].border = b_data_cell

                # TOTAL 행 추가
                total_row_idx = start_row + total_data_count
                b_total_cell = Border(top=thin_side, bottom=med_side, left=thin_side, right=thin_side)

                ws[f"E{total_row_idx}"] = "Total"
                ws[f"E{total_row_idx}"].font = font_calibri_bold
                ws[f"E{total_row_idx}"].border = b_total_cell

                formatted_total_kgs = format_number(sum_kgs)
                try:
                    ws[f"F{total_row_idx}"] = float(formatted_total_kgs)
                except:
                    ws[f"F{total_row_idx}"] = formatted_total_kgs
                ws[f"F{total_row_idx}"].font = font_calibri_bold
                ws[f"F{total_row_idx}"].alignment = align_center
                ws[f"F{total_row_idx}"].border = b_total_cell

                ws[f"G{total_row_idx}"] = sum_pkgs
                ws[f"G{total_row_idx}"].font = font_calibri_bold
                ws[f"G{total_row_idx}"].alignment = align_center
                ws[f"G{total_row_idx}"].border = b_total_cell

                ws[f"H{total_row_idx}"] = "PKG'S"
                ws[f"H{total_row_idx}"].font = font_calibri_bold
                ws[f"H{total_row_idx}"].alignment = align_center
                ws[f"H{total_row_idx}"].border = b_total_cell

                formatted_total_cbm = format_number(sum_cbm)
                try:
                    ws[f"I{total_row_idx}"] = float(formatted_total_cbm)
                except:
                    ws[f"I{total_row_idx}"] = formatted_total_cbm
                ws[f"I{total_row_idx}"].font = font_calibri_bold
                ws[f"I{total_row_idx}"].alignment = align_center
                ws[f"I{total_row_idx}"].border = b_total_cell

                col_widths = {
                    'A': 8.38, 'B': 15.38, 'C': 9.38, 'D': 40.13, 'E': 54.13,
                    'F': 13.0, 'G': 10.75, 'H': 11.63, 'I': 10.75, 'J': 10.75,
                    'K': 13.63, 'L': 13.88, 'M': 17.75
                }
                for c_letter, w_val in col_widths.items():
                    ws.column_dimensions[c_letter].width = w_val

                output_excel = io.BytesIO()
                wb.save(output_excel)
                output_excel.seek(0)

                st.write("")
                st.download_button(
                    label="💾 IST FINAL CONSOL LIST 다운로드",
                    data=output_excel,
                    file_name=f"IST_FINAL_CONSOL_LIST_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:
                st.error("엑셀 파일 내 'House B/L No' 열을 찾을 수 없습니다.")

        except Exception as e:
            st.error(f"IST 데이터 처리 중 오류 발생: {e}")

# ==========================================
# TAB 4: 선적이력
# ==========================================
with tab_history:
    col_pod, col_query, col_btn = st.columns([1, 1.8, 0.5])
    with col_pod:
        selected_pod_opt = st.selectbox("POD 선택", POD_OPTIONS, index=0)
        
        if "ALL" in selected_pod_opt:
            pod_code = "ALL"
        else:
            pod_code = selected_pod_opt.split("(")[-1].replace(")", "").strip()
            
    with col_query:
        search_query = st.text_input("HS CODE 또는 품목명 검색", placeholder="예: AUTOMOTIVE, 844391, 2008.99, MAGNET 등")
    with col_btn:
        st.write("") 
        st.write("")
        search_btn = st.button("🔍 검색", use_container_width=True)
    
    st.divider()

    if search_query.strip():
        search_upper = search_query.strip().upper()
        search_digits = re.sub(r'[^0-9]', '', search_upper)
        history_warnings = []

        if "200899" in search_digits or "2008.99" in search_upper:
            history_warnings.append("⚠️ 2008.99-9000 EU 관세 품목 분류에 등록 되지 않은 코드")
        if "242400" in search_digits or "2424.00" in search_upper:
            history_warnings.append("⚠️ 유효하지 않은 HS CODE / HOUSEHOLD GOODS 는 9905.00 을 써주세요.")
        if "MAGNET" in search_upper or "자성" in search_upper:
            history_warnings.append("⚠️ 자성물질 MSDS 필요")
        if "CARBON" in search_upper or "카본" in search_upper:
            history_warnings.append("⚠️ carbon DG 로 간주되어 LCL 선적 불가")
        if "ALKALINE" in search_upper or "알카라인" in search_upper:
            history_warnings.append("ℹ️ Alkaline Battery 선적 가능")
        if "LITHIUM" in search_upper or "리튬" in search_upper:
            history_warnings.append("⚠️ lithium battery NON-DG 여도 LCL 선적 불가")
        if any(kw in search_upper for kw in ["FOOD", "식품", "음식"]):
            history_warnings.append("⚠️ FOOD STUFF 도착지 확인 필요")

        if history_warnings:
            combined_hist_warning = "\n".join(history_warnings)
            st.markdown(f'<div style="display:inline-block;padding:5px 15px;border-radius:5px;background-color:rgba(255, 75, 75, 0.1);border:1px solid rgb(255, 75, 75);color:rgb(255, 75, 75);font-family:sans-serif;font-size:14px;line-height:1.5;white-space:pre-wrap;margin-bottom:15px;">{combined_hist_warning}</div><br>', unsafe_allow_html=True)

    files_to_scan = []
    if pod_code == "ALL":
        for country, code in POD_LIST:
            if code == "ALL": continue
            pfile = f"{code}.xlsx"
            if os.path.exists(pfile):
                files_to_scan.append((code, pfile))
            elif code == "PLGDN" and os.path.exists("GDN.xlsx"):
                files_to_scan.append((code, "GDN.xlsx"))
    else:
        pfile = f"{pod_code}.xlsx"
        if os.path.exists(pfile):
            files_to_scan.append((pod_code, pfile))
        elif pod_code == "PLGDN" and os.path.exists("GDN.xlsx"):
            files_to_scan.append((pod_code, "GDN.xlsx"))
            
    if files_to_scan:
        all_matched_rows = []
        for cur_pod, target_file in files_to_scan:
            try:
                raw_df = pd.read_excel(target_file)
                
                header_row_idx = None
                if "House B/L No" in raw_df.columns:
                    hist_df = raw_df
                else:
                    for idx in range(min(5, len(raw_df))):
                        row_vals = [str(v).strip() for v in raw_df.iloc[idx].values]
                        if any("House B/L" in v for v in row_vals):
                            header_row_idx = idx
                            break
                    if header_row_idx is not None:
                        hist_df = pd.read_excel(target_file, header=header_row_idx + 1)
                    else:
                        hist_df = raw_df

                hist_df.columns = [str(c).strip() for c in hist_df.columns]
                
                hbl_col = next((c for c in hist_df.columns if "House B/L" in c), None)
                etd_col = next((c for c in hist_df.columns if "ETD" in c), None)
                item_col = next((c for c in hist_df.columns if "품목" in c), None)
                
                if hbl_col and etd_col and item_col:
                    res_df = hist_df[[hbl_col, etd_col, item_col]].copy()
                    res_df.columns = ['House B/L No', 'ETD', '품목']
                    res_df = res_df.dropna(subset=['House B/L No'])
                    
                    if search_query.strip():
                        q_raw = search_query.strip().upper()
                        q_digits = re.sub(r'[^0-9]', '', q_raw)
                        
                        for _, r in res_df.iterrows():
                            item_val = str(r['품목']) if pd.notna(r['품목']) else ""
                            item_upper = item_val.upper()
                            item_digits = re.sub(r'[^0-9]', '', item_val)
                            
                            is_match = False
                            if len(q_digits) >= 4 and q_digits in item_digits:
                                is_match = True
                            elif q_raw in item_upper:
                                is_match = True
                                
                            if is_match:
                                etd_str = str(r['ETD']).split(' ')[0] if pd.notna(r['ETD']) else ""
                                item_dict = {
                                    'House B/L No': str(r['House B/L No']).strip(),
                                    'ETD': etd_str,
                                    '품목': item_val.strip()
                                }
                                if pod_code == "ALL":
                                    item_dict = {'POD': cur_pod, **item_dict}
                                all_matched_rows.append(item_dict)
            except Exception as e:
                pass
                
        if search_query.strip():
            if all_matched_rows:
                out_df = pd.DataFrame(all_matched_rows)
                st.subheader(f"🔍 검색 결과 ({len(out_df)}건)")
                
                cfg = {
                    "House B/L No": st.column_config.TextColumn("House B/L No", width=160),
                    "ETD": st.column_config.TextColumn("ETD", width=110),
                    "품목": st.column_config.TextColumn("품목", width=400),
                }
                if pod_code == "ALL":
                    cfg = {"POD": st.column_config.TextColumn("POD", width=90), **cfg}
                    
                st.dataframe(
                    out_df,
                    column_config=cfg,
                    use_container_width=False,
                    height=500
                )
            else:
                st.info("검색 조건에 맞는 이력이 없습니다.")
        else:
            st.write("💡 HS CODE 또는 품목명을 입력 후 Enter를 누르거나 [🔍 검색] 버튼을 누르면 해당 POD의 진행 이력을 검색합니다.")
    else:
        st.warning("저장된 이력 엑셀 파일이 없습니다. (루트 폴더에 포트코드.xlsx 파일을 넣어주세요)")

# ==========================================
# TAB 5: 업로드 기록
# ==========================================
with tab2:
    if os.path.exists("upload_log.txt"):
        with open("upload_log.txt", "r", encoding='utf-8') as f: 
            st.text_area("Log", f.read(), height=800)
